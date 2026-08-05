"""Парсер каталога товаров в таблицу Excel.

Демонстрационный сайт — books.toscrape.com, песочница, специально созданная
для обучения парсингу. Под конкретный заказ меняются только селекторы
в extract_page и колонки в COLUMNS.
"""
from __future__ import annotations

import argparse
import re
import time
from dataclasses import astuple, dataclass, fields
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = "https://books.toscrape.com/catalogue/page-{}.html"
RATINGS = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}
HEADERS = {"User-Agent": "portfolio-demo-parser/1.0"}


@dataclass
class Item:
    title: str
    price: float
    rating: int
    in_stock: bool
    url: str


COLUMNS = ["Название", "Цена, £", "Рейтинг", "В наличии", "Ссылка"]


def fetch(url: str, retries: int = 3) -> str:
    """Забирает страницу, повторяя при сетевых сбоях с нарастающей паузой."""
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            r.encoding = r.apparent_encoding
            return r.text
        except requests.RequestException as e:
            if attempt == retries:
                raise
            print(f"  сбой ({e}), повтор {attempt}/{retries - 1}")
            time.sleep(attempt * 2)
    raise RuntimeError("недостижимо")


def extract_page(html: str) -> list[Item]:
    soup = BeautifulSoup(html, "html.parser")
    items: list[Item] = []
    for pod in soup.select("article.product_pod"):
        link = pod.select_one("h3 a")
        price_text = pod.select_one("p.price_color").get_text(strip=True)
        rating_class = pod.select_one("p.star-rating")["class"]
        rating = next((RATINGS[c] for c in rating_class if c in RATINGS), 0)
        items.append(
            Item(
                title=link["title"].strip(),
                # цена уходит числом, а не текстом: заказчик сможет сортировать и суммировать
                price=float(re.sub(r"[^\d.]", "", price_text)),
                rating=rating,
                in_stock="in stock" in pod.select_one("p.instock").get_text().lower(),
                url="https://books.toscrape.com/catalogue/" + link["href"],
            )
        )
    return items


def save_xlsx(items: list[Item], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"

    ws.append(COLUMNS)
    head_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")

    for item in items:
        row = list(astuple(item))
        row[3] = "да" if row[3] else "нет"
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [58, 11, 10, 12, 62]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "0.00"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser(description="Парсер каталога в xlsx")
    ap.add_argument("--pages", type=int, default=50, help="сколько страниц забрать")
    ap.add_argument("--delay", type=float, default=0.4, help="пауза между запросами, сек")
    ap.add_argument("--out", type=Path, default=Path("output/catalog.xlsx"))
    args = ap.parse_args()

    started = time.monotonic()
    items: list[Item] = []
    for page in range(1, args.pages + 1):
        print(f"страница {page}/{args.pages}", end="\r", flush=True)
        page_items = extract_page(fetch(BASE.format(page)))
        if not page_items:
            break
        items.extend(page_items)
        time.sleep(args.delay)  # не долбим чужой сервер

    save_xlsx(items, args.out)
    took = time.monotonic() - started
    print(f"\nсобрано {len(items)} товаров за {took:.0f} с → {args.out}")


if __name__ == "__main__":
    main()
