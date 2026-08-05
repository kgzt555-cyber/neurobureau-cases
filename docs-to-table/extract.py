"""Разбор пачки счетов PDF в таблицу Excel.

Текст из PDF вытаскивается локально, разбирает его языковая модель:
у счетов от разных поставщиков всегда разная вёрстка, и регулярки на этом ломаются.
"""
from __future__ import annotations

import argparse
import json
import os
import time
from pathlib import Path

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

HERE = Path(__file__).parent
load_dotenv(HERE.parent / ".env")

PROMPT = """Ты разбираешь счёт на оплату. Верни СТРОГО JSON без пояснений:

{
  "number": "номер счёта как в документе",
  "date": "дата в формате ДД.ММ.ГГГГ",
  "supplier": "название поставщика",
  "inn": "ИНН строкой",
  "vat_rate": число процентов НДС (0 если не облагается),
  "items": [{"name": "наименование", "qty": число, "price": число, "sum": число}],
  "subtotal": число (итого без НДС),
  "vat": число (сумма НДС),
  "total": число (всего к оплате)
}

Числа — без пробелов и знака валюты, дробная часть через точку.
Если поля нет в документе, ставь null. Ничего не выдумывай.

Текст счёта:
"""


def pdf_text(path: Path) -> str:
    return "\n".join(page.extract_text() or "" for page in PdfReader(path).pages)


def parse_invoice(client: OpenAI, model: str, text: str) -> dict:
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": PROMPT + text}],
        response_format={"type": "json_object"},
        temperature=0,
    )
    return json.loads(resp.choices[0].message.content)


def save_xlsx(invoices: list[dict], path: Path) -> None:
    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1F2937")

    def style_header(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    ws = wb.active
    ws.title = "Счета"
    ws.append(["Файл", "Номер", "Дата", "Поставщик", "ИНН",
               "НДС, %", "Без НДС", "НДС", "Всего"])
    for inv in invoices:
        ws.append([inv.get("file"), inv.get("number"), inv.get("date"),
                   inv.get("supplier"), inv.get("inn"), inv.get("vat_rate"),
                   inv.get("subtotal"), inv.get("vat"), inv.get("total")])
    style_header(ws)
    for w, i in zip([22, 16, 12, 26, 15, 9, 13, 12, 13], range(1, 10)):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=9):
        for cell in row:
            cell.number_format = "# ##0.00"

    ws2 = wb.create_sheet("Позиции")
    ws2.append(["Счёт", "Наименование", "Кол-во", "Цена", "Сумма"])
    for inv in invoices:
        for item in inv.get("items") or []:
            ws2.append([inv.get("number"), item.get("name"),
                        item.get("qty"), item.get("price"), item.get("sum")])
    style_header(ws2)
    for w, i in zip([16, 44, 10, 13, 14], range(1, 6)):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "# ##0.00"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser(description="Счета PDF в таблицу Excel")
    ap.add_argument("--src", type=Path, default=HERE / "samples")
    ap.add_argument("--out", type=Path, default=HERE / "output" / "invoices.xlsx")
    args = ap.parse_args()

    client = OpenAI(
        api_key=os.environ["AI_API_KEY"],
        base_url=os.environ.get("AI_BASE_URL", "https://api.proxyapi.ru/openai/v1"),
    )
    model = os.environ.get("AI_MODEL", "gpt-4o-mini")

    started = time.monotonic()
    invoices = []
    files = sorted(args.src.glob("*.pdf"))
    for n, pdf in enumerate(files, start=1):
        print(f"[{n}/{len(files)}] {pdf.name}")
        data = parse_invoice(client, model, pdf_text(pdf))
        data["file"] = pdf.name
        invoices.append(data)

    save_xlsx(invoices, args.out)
    (args.out.parent / "parsed.json").write_text(
        json.dumps(invoices, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    took = time.monotonic() - started
    print(f"\nразобрано {len(invoices)} счетов за {took:.0f} с → {args.out}")


if __name__ == "__main__":
    main()
